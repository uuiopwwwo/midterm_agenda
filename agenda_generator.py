from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 议程数据
agenda_data = [
    ("09:00~09:30", "報到", "無"),
    ("09:30~09:50", "開場致詞", "林院長"),
    ("09:50~10:20", "5G 專網與企業數位轉型的實踐之路", "張教授"),
    ("10:20~10:50", "AI 驅動的智慧工廠無線網路架構設計", "李教授"),
    ("10:50~11:10", "Break", "無"),
    ("11:10~11:40", "Open RAN 生態系統的發展與挑戰", "研究團隊"),
    ("11:40~12:10", "雲原生架構在 5G 核心網路中的應用", "研究團隊"),
    ("12:10~13:30", "Lunch", "無"),
    ("13:30~14:00", "毫米波通訊技術與室內定位整合應用", "研究團隊"),
    ("14:00~14:30", "網路切片技術於醫療場域的實證研究", "胡教授"),
    ("14:30~14:50", "Break", "無"),
    ("14:50~15:30", "O-RAN 近即時控制器 xApp 開發實戰", "研究團隊"),
    ("15:30~16:00", "綜合座談與交流", "研究團隊"),
    ("16:00~16:20", "Break", "無"),
    ("16:20~17:00", "Energy-Efficient Resource Allocation in O-RAN Architecture", "賴博士"),
    ("17:00~17:40", "Deep Reinforcement Learning for Network Slicing Optimization", "賴博士"),
    ("17:40~18:10", "Dinner", "無"),
    ("18:10~18:50", "Federated Learning Approaches for Privacy-Preserving 6G Networks", "陳教授"),
    ("18:50~19:30", "Digital Twin-Enabled Intelligent RAN Management", "陳教授"),
    ("19:30~20:00", "Panel Discussion and Closing Remarks", "研究團隊"),
]

# 創建工作簿
wb = Workbook()
ws = wb.active
ws.title = "會議議程"

# 定義樣式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
wrap_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 設定列寬
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 15

# 添加標題行
headers = ["時間", "內容", "講者"]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = border

# 添加數據
for row_num, (time, content, speaker) in enumerate(agenda_data, 2):
    # 時間列
    time_cell = ws.cell(row=row_num, column=1)
    time_cell.value = time
    time_cell.alignment = center_alignment
    time_cell.border = border
    
    # 內容列
    content_cell = ws.cell(row=row_num, column=2)
    content_cell.value = content
    content_cell.alignment = wrap_alignment
    content_cell.border = border
    
    # 講者列
    speaker_cell = ws.cell(row=row_num, column=3)
    speaker_cell.value = speaker
    speaker_cell.alignment = center_alignment
    speaker_cell.border = border

# 設定行高
ws.row_dimensions[1].height = 25
for row_num in range(2, len(agenda_data) + 2):
    ws.row_dimensions[row_num].height = 30

# 保存文件
output_file = "/home/student/midterm_agenda/會議議程.xlsx"
wb.save(output_file)
print(f"Excel 文件已成功生成：{output_file}")
print(f"總共 {len(agenda_data)} 條議程項目")
